import logging
from io import BytesIO
import openpyxl
from ..db_operations import insert_row, create_table, is_table_dynamic, dynamic_mapping_columns
from .sanitize_names import sanitize_field_names, sanitize_table_name

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def find_longest_row(ws):
    longest_row_length = 0
    for row in ws.iter_rows(values_only=True):
        row_length = len(row)
        while row_length > 0 and row[row_length - 1] is None:
            row_length -= 1
        if row_length > longest_row_length:
            longest_row_length = row_length
    return longest_row_length


def detect_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        non_empty_count = len([cell for cell in row if cell is not None])
        string_count = sum([1 for cell in row if isinstance(cell, str)])
        if non_empty_count > 2 and string_count / non_empty_count > 0.5:
            return i
    return None


def split_and_insert_dynamic_data(sheet_table_name, field_names, values, static_column_names, dynamic_column_names):
    static_data = {field: value for field, value in zip(field_names, values) if field in static_column_names}
    dynamic_data = {field: value for field, value in zip(field_names, values) if field in dynamic_column_names}

    static_insert_data = {
        "schema": "data",
        "name": sheet_table_name,
        "field_names": list(static_data.keys()),
        "data": list(static_data.values())
    }
    insert_row(static_insert_data)

    dynamic_insert_data = {
        "schema": "data",
        "name": sheet_table_name,
        "field_names": list(dynamic_data.keys()),
        "data": list(dynamic_data.values())
    }
    insert_row(dynamic_insert_data)


def process_rows(ws, header_row_index, field_names, sheet_table_name, is_dynamic, static_column_names, dynamic_column_names):
    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        if not any(row):
            continue

        values = [str(value) if value is not None else '' for value in row]
        if len(values) != len(field_names):
            raise ValueError("Number of values in row doesn't match number of fields")

        if is_dynamic:
            split_and_insert_dynamic_data(sheet_table_name, field_names, values, static_column_names, dynamic_column_names)
        else:
            insert_data = {
                "schema": "data",
                "name": sheet_table_name,
                "field_names": field_names,
                "data": values
            }
            insert_row(insert_data)


def check_dynamic_and_get_columns(sheet_table_name):
    is_dynamic = is_table_dynamic(sheet_table_name)
    if is_dynamic:
        static_column_names, dynamic_column_names = dynamic_mapping_columns(sheet_table_name, "data")
        if not static_column_names or not dynamic_column_names:
            is_dynamic = False
    else:
        static_column_names, dynamic_column_names = None, None
    return is_dynamic, static_column_names, dynamic_column_names


def handle_special_keywords(ws):
    longest_row = find_longest_row(ws)
    if longest_row is None:
        logger.warning("No header row detected in sheet. Skipping this sheet.")
        return [], None
    header_row_index = 0
    field_names = [f"column_{i}" for i in range(longest_row)]
    return field_names, header_row_index


def handle_standard_keywords(ws):
    header_row_index = detect_header_row(ws)
    if header_row_index is None:
        logger.warning("No header row detected in sheet. Skipping this sheet.")
        return [], None
    header_row_index += 1
    field_names = sanitize_field_names([cell.value for cell in ws[header_row_index]])
    return field_names, header_row_index


def process_sheet(wb, sheet_name, id, table_name):
    ws = wb[sheet_name]
    keywords = ["param", "def", "déf", "conf", "doc"]

    if any(keyword in sheet_name.lower() for keyword in keywords):
        field_names, header_row_index = handle_special_keywords(ws)
    else:
        field_names, header_row_index = handle_standard_keywords(ws)

    sheet_table_name = f"{table_name}_{sanitize_table_name(sheet_name)}"

    table_data = {
        "schema": "data",
        "name": sheet_table_name,
        "field_names": field_names,
        "id": id
    }

    try:
        create_table(table_data)
    except Exception as create_table_error:
        logger.error(f"Error creating table {sheet_table_name}: {create_table_error}")
        return

    is_dynamic, static_column_names, dynamic_column_names = check_dynamic_and_get_columns(sheet_table_name)
    
    try:
        process_rows(ws, header_row_index, field_names, sheet_table_name, is_dynamic, static_column_names, dynamic_column_names)
    except Exception as create_table_error:
        logger.error(f"Error uploading the data: {create_table_error}")



def process_xlsx(id, data, table_name):
    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
        for sheet_name in wb.sheetnames:
            process_sheet(wb, sheet_name, id, table_name)
    except Exception as xlsx_process_error:
        logger.error(f"Error processing XLSX: {xlsx_process_error}")
        raise xlsx_process_error
